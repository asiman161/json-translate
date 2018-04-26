"""

All *.py files combined into one. Maybe deprecated.

merge ['excel_parser.py', 'json_parser.py', 'list_to_json.py', 'translate.py'].
If this file works bad use 'translate.py' instead of this.

"""
import json

import sys
from openpyxl import load_workbook


def create_words_dict(sheet, offset_row=0, column_word_index=0, column_word_offset=0):
    words_dict = {}

    for i, row in enumerate(sheet.rows):
        if sheet.cell(row=i + offset_row, column=column_word_index).value is not None:
            words_dict[sheet.cell(row=i + offset_row, column=column_word_index).value.strip().lower()] \
                = sheet.cell(row=i + offset_row, column=column_word_index + column_word_offset).value.strip()

    return words_dict


def parse(data, keys, index, keys_by_translate):
    for i, key in enumerate(data):
        key = key
        k = keys[:index]
        if isinstance(data[key], dict):
            if len(k) <= index:
                k.append(key)
            else:
                k[index] = key
            parse(data[key], k, index + 1, keys_by_translate)
        elif isinstance(data[key], list):
            k.append(key)
            for index_inner, data_key in enumerate(data[key]):
                last_key = k[len(k) - 1]
                if isinstance(last_key, int) and last_key + 1 == index_inner:
                    k[len(k) - 1] = index_inner
                else:
                    k.append(index_inner)
                parse(data[key][index_inner], k, index + 2, keys_by_translate)
        elif not isinstance(data[key], bool):
            if len(k) == index:
                k.append(key)
            elif index > len(k):
                k.append(key)
            else:
                k[index] = key
            if data[key] in keys_by_translate:
                keys_by_translate[data[key].strip()].append(k)
            else:
                keys_by_translate[data[key].strip()] = [k]


def write_exists(exists, dict_exists, write_exist=False):
    from openpyxl import Workbook
    index = 2

    book = Workbook()

    sheet_local = book.active
    sheet_local['C1'] = "Original"
    sheet_local['D1'] = "Translate"

    for i, text in enumerate(exists[False]):
        sheet_local['C' + str(index)] = text
        sheet_local['D' + str(index)] = "-"
        index += 1

    # print translates that exists
    if write_exist:
        index += 1
        for text in exists[True]:
            sheet_local['C' + str(index)] = text
            sheet_local['D' + str(index)] = dict_exists[text.lower()]
            index += 1

    book.save('all.xlsx')


def switch_keys(old_dict, excel_dict, log=False, write_exist=False):
    copy = {}
    exists = {True: [], False: []}
    for i, key in enumerate(old_dict):
        # TODO: regex to find that file contains interpolation `{{...}}`: (\{\{)([\w\s_]*)(\}\})
        if key.lower() in excel_dict:
            if excel_dict[key.lower()] in copy:
                copy[excel_dict[key.lower()]].append(old_dict[key][0])
            else:
                copy[excel_dict[key.lower()]] = old_dict[key]
            exists[True].append(key)
        else:
            exists[False].append(key)
            copy[key] = old_dict[key]
    write_exists(exists, excel_dict, write_exist)
    if log:
        print('translates don\'t exist: {len_not_exist}\ntranslates exist: {len_exist}\n{not_exist}'.format(
            len_not_exist=len(exists[False]),
            len_exist=len(exists[True]),
            not_exist=exists[False]
        ))
    return copy


def create_deep_dict(value, layers):
    orig_data = {}
    data = orig_data
    last_layer = layers[-1]

    for layer in layers[:-1]:
        data[layer] = {}
        data = data[layer]

    data[last_layer] = value

    return orig_data


def fill_arr(data):
    arr_from_dict = []
    for key in data:
        if len(data[key]) > 1:
            for i, inner_arr in enumerate(data[key]):
                arr_from_dict.append({key: inner_arr})
        else:
            arr_from_dict.append({key: data[key][0]})

    return arr_from_dict


def transform_list_to_dict(data):
    items = [{}] * len(data)
    for i, item in enumerate(data):
        for key in item:
            items[i] = create_deep_dict(key, item[key])
    return items


def merge(a, b, path=None):
    if path is None:
        path = []
    for key in b:
        if key in a:
            if isinstance(a[key], dict) and isinstance(b[key], dict):
                merge(a[key], b[key], path + [str(key)])
            elif a[key] == b[key]:
                pass
            else:
                raise Exception('Conflict at %s' % '.'.join(path + [str(key)]))
        else:
            a[key] = b[key]
    return a


if __name__ == '__main__':
    from_translate, to_translate, dictionary, log, write_exist = None, None, None, False, False
    for i, arg in enumerate(sys.argv):
        if arg == '--from':
            from_translate = sys.argv[i + 1]
        if arg == '--to':
            to_translate = sys.argv[i + 1]
        if arg == '--dict':
            dictionary = sys.argv[i + 1]
        if arg == '--log' and sys.argv[i + 1].lower() == 'true':
            log = True
        if arg == '--write-exist' and sys.argv[i + 1].lower() == 'true':
            write_exist = True
    if from_translate is None or to_translate is None or dictionary is None:
        raise Exception('''not enough arguments.
            should be path to file from, file to and dictionary with translates''')
    keys_by_translate = {}
    with open(from_translate, 'r') as fp:
        parsed = json.load(fp)
        parse(parsed, [], 0, keys_by_translate)

        wb = load_workbook(dictionary)
        wb.close()
        sheet = wb[wb.sheetnames[0]]
        words = create_words_dict(sheet, offset_row=2, column_word_index=1, column_word_offset=1)
        keys_by_translate = switch_keys(keys_by_translate, words, log=log, write_exist=write_exist)

    arr = fill_arr(keys_by_translate)
    arr = transform_list_to_dict(arr)

    cor_obj = arr[0]
    for obj in arr[1:]:
        cor_obj = merge(cor_obj, obj)

    with open(to_translate, 'w') as fp:
        json.dump(cor_obj, fp, ensure_ascii=False, indent=4)
