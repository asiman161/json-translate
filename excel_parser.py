def create_words_dict(sheet, offset_row=0, column=0, column_offset=0):
    """Создает объект, в котором хранится изначальная фраза как ключ и ее перевод как значение
    :arg sheet: xlsx файл
    :arg offset_row: сколько строк необходимо пропустить
    :arg column: на какой колонке находятся изначальные переводы
    :arg column_offset: с какиим смещением находятся переводы"""
    words_dict = {}

    for i, row in enumerate(sheet.rows):
        if sheet.cell(row=i + offset_row, column=column).value is not None:
            words_dict[sheet.cell(row=i + offset_row, column=column).value.strip().lower()] \
                = sheet.cell(row=i + offset_row, column=column + column_offset).value.strip()

    return words_dict


if __name__ == '__main__':
    from openpyxl import load_workbook

    wb = load_workbook('Interface_translate_MN_29.xlsx')
    sheet = wb[wb.sheetnames[0]]

    words = create_words_dict(sheet, offset_row=2, column=3, column_offset=1)
    wb.close()
    print(words)
