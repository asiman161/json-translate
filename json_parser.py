def parse(data, keys, index, keys_by_translate):
    for i, key in enumerate(data):
        key = key
        k = keys[:index]
        if isinstance(data[key], dict):
            if len(k) <= index:
                k.append(key)
            else:
                k[index] = key
            parse(data[key], k, index + 1, keys_by_translate)
        elif isinstance(data[key], list):
            k.append(key)
            for index_inner, data_key in enumerate(data[key]):
                last_key = k[len(k) - 1]
                if isinstance(last_key, int) and last_key + 1 == index_inner:
                    k[len(k) - 1] = index_inner
                else:
                    k.append(index_inner)
                parse(data[key][index_inner], k, index + 2, keys_by_translate)
        elif not isinstance(data[key], bool):
            if len(k) == index:
                k.append(key)
            elif index > len(k):
                k.append(key)
            else:
                k[index] = key
            if data[key] in keys_by_translate:
                keys_by_translate[data[key].strip()].append(k)
            else:
                keys_by_translate[data[key].strip()] = [k]


def write_exists(exists, dict_exists, write_exist=False):
    from openpyxl import Workbook
    index = 2

    book = Workbook()

    sheet_local = book.active
    sheet_local['C1'] = "Original"
    sheet_local['D1'] = "Translate"

    for i, text in enumerate(exists[False]):
        sheet_local['C' + str(index)] = text
        sheet_local['D' + str(index)] = "-"
        index += 1

    # print translates that exists
    if write_exist:
        index += 1
        for text in exists[True]:
            sheet_local['C' + str(index)] = text
            sheet_local['D' + str(index)] = dict_exists[text.lower()]
            index += 1

    book.save('all.xlsx')


def switch_keys(old_dict, excel_dict, log=False, write_exist=False):
    copy = {}
    exists = {True: [], False: []}
    for i, key in enumerate(old_dict):
        # TODO: regex to find that file contains interpolation `{{...}}`: (\{\{)([\w\s_]*)(\}\})
        if key.lower() in excel_dict:
            if excel_dict[key.lower()] in copy:
                copy[excel_dict[key.lower()]].append(old_dict[key][0])
            else:
                copy[excel_dict[key.lower()]] = old_dict[key]
            exists[True].append(key)
        else:
            exists[False].append(key)
            copy[key] = old_dict[key]
    write_exists(exists, excel_dict, write_exist)
    if log:
        print('translates don\'t exist: {len_not_exist}\ntranslates exist: {len_exist}\n{not_exist}'.format(
            len_not_exist=len(exists[False]),
            len_exist=len(exists[True]),
            not_exist=exists[False]
        ))
    return copy


if __name__ == '__main__':
    import json

    from openpyxl import load_workbook

    from excel_parser import create_words_dict

    parsed = json.load(open('translates/ru.json'))

    all_keys = {}
    parse(parsed, [], 0, all_keys)
    wb = load_workbook('Interface_translate_MN_29.xlsx')
    sheet = wb[wb.sheetnames[0]]
    words = create_words_dict(sheet, offset_row=2, column_word_index=3, column_word_offset=1)
    new_all_keys = switch_keys(all_keys, words, log=True)

    print(all_keys)
    print(new_all_keys)
