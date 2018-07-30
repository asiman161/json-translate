"""
Генерирует JSON файл перевода, основываясь на xlsx таблице,
в которой попарно хранится слово или фраза на одном языке и его перевод на другой
и иисходном JSON файле, на основаниии которого генерируется новый
"""
import json
import sys

from openpyxl import load_workbook

from excel_parser import create_words_dict
from json_parser import parse, switch_keys
from list_to_json import fill_arr, transform_list_to_dict, merge

if __name__ == '__main__':
    """
    Делает перевод на основе исходного JSON файла для генерации нового и словаря из xlsx файла
    :flags:
        --from - JSON файл, на основе которого будет сгенерирован новый JSON файл
        --to - JSON файл, который будет сгенерирован
        --dict - xlsx файл, в нем хранятся попарные переводы
        --log - выведет лог в консоль о переводах, где будет указано кол-во существующих переводов, кол-во отсутствующих 
        и их список 
        --write-exist - добавит в xlsx файл переводы тех слов, которые уже существуют и те, которых еще нет. 
        имя файла захардкожено как all.xlsx 
    """
    from_translate, to_translate, dictionary, log, write_exist = None, 'lang.json', None, False, True
    for i, arg in enumerate(sys.argv):
        if arg == '--from' and sys.argv[i + 1] is not None:
            from_translate = sys.argv[i + 1]
        if arg == '--to' and sys.argv[i + 1] is not None:
            to_translate = sys.argv[i + 1]
        if arg == '--dict' and sys.argv[i + 1] is not None:
            dictionary = sys.argv[i + 1]
        if arg == '--log':
            log = True
        if arg == '--write-exist':
            write_exist = True
    if from_translate is None or dictionary is None:
        raise Exception('''not enough arguments. Should be path to file from and dictionary with translates''')

    """переводы исходного JSON'a. В нем перевод выступает в качестве ключа для ключей"""
    keys_by_translate = {}
    with open(from_translate, 'r') as fp:
        """спаршенный исходный JSON"""
        parsed = json.load(fp)

        """JSON, вывернутый наизнанку"""
        parse(parsed, [], 0, keys_by_translate)

        wb = load_workbook(dictionary)
        wb.close()
        """какая страница xlsx будет открыта. Хардкод и требует постоянного изменения"""
        sheet = wb[wb.sheetnames[1]]

        """объект, в котором хранятся слова и их переводы. 
        Слова приводятся к нижнему регистру с целью снижения конфликтов
        :offset_row, :column, :column_offset - требуется постоянная перезапись в зависимости от исходного xlsx файла
        """
        words = create_words_dict(sheet, offset_row=2, column=1, column_offset=1)

        """исходные слова заменены на новые. 
        Теперь в кач-ве ключей используются слова из второго(целевого) языка."""
        keys_by_translate = switch_keys(keys_by_translate, words, log=log, write_exist=write_exist)

    """Перевод в массив объектов"""
    arr = fill_arr(keys_by_translate)
    """содержит массив объектов. Переводы и ключи поменялись местами.
    Теперь ключи выполняют роль ключей и переводы в роли значений"""
    arr = transform_list_to_dict(arr)

    """Объединение массива объектов в один большой. Он же является целевым объектом
    для дампа"""
    cor_obj = arr[0]
    for obj in arr[1:]:
        cor_obj = merge(cor_obj, obj)

    """Сохранение проделанной работы"""
    with open(to_translate, 'w') as fp:
        json.dump(cor_obj, fp, ensure_ascii=False, indent=4)
